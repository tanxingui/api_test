#encoding=utf-8
from openpyxl import load_workbook
from common.MyException import exception_utils
from common.BaseApi import BaseApi
from checkout.BaseRemind import Remind
from common.BaseDriver import BaseDriver
from config.mypath import *

MyRemind=Remind()
MyDriver=BaseDriver()
@exception_utils
class ExcelUtil(BaseApi):
    """封装操作excel的方法，主要作用2个：
1、用于数据处理，读取excel中用例后返回规定的字典，便于生成yaml
2、将运行结果数据存入excel中对应列
"""
    def __init__(self, excel_path):
        self.wb = load_workbook(excel_path)
        self.template = """{"case_name":"","precondition":"","step":"","request_list":"","assert":""},"""  # 这个是写入用例的模板
        self.common_template="""{"id":0,"api_name":"","url":"","method":"","body":"","is_replace":""},"""  # 这个是公共用例的模板


    @exception_utils
    def read_excel(self):
        """读取excel，处理数据，并返回一个格式处理后的字典"""
        value = []
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            case_list_excel = list(ws.values)
            case_list = MyRemind.Excel_manage(case_list_excel)
            cases_template = self.template * len(case_list)
            cases_template_list = eval("[" + cases_template[:-1] + "]")   # 与用例相同长度的模板
            common_value=ExcelUtil(excel_dir+"/common.xlsx").read_common()

            for i in range(len(case_list)):  # i：第i个用例
                # 每个用例中字段是5个，因此这样写
                url_list=case_list[i][3].split(",")
                step=case_list[i][2]
                request_list=MyDriver.url_list_driver(url_list,step,common_value)
                cases_template_list[i]['case_name'] = case_list[i][0]
                cases_template_list[i]['precondition'] = case_list[i][1]
                cases_template_list[i]['step'] = case_list[i][2]
                cases_template_list[i]['request_list'] = request_list
                cases_template_list[i]['assert'] = case_list[i][4]

            value.append({"cases": cases_template_list})

        return value




    @exception_utils
    def read_common(self):
        common_value={}
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            case_list_excel = list(ws.values)
            api_list = MyRemind.Common_Excel_manage(case_list_excel)  #格式处理器 变更模板需要变更格式处理器
            cases_template = self.common_template * len(api_list)  # 一个sheet中用例的数量
            cases_template_list = eval("[" + cases_template[:-1] + "]")

            # aaa=[{"url":{"url":1,"name":"lu"}}]
            for i in range(len(api_list)):  # i：第i个接口
                # 每个用例中字段是6个，因此这样写
                cases_template_list[i]['id'] = api_list[i][0]
                cases_template_list[i]['api_name'] = api_list[i][1]
                cases_template_list[i]['url'] = api_list[i][2]
                cases_template_list[i]['method'] = api_list[i][3]
                cases_template_list[i]['body'] = api_list[i][4]
                cases_template_list[i]['is_replace'] = api_list[i][5]

                common_value.update({f"{cases_template_list[i]['url']}": cases_template_list[i]})

            return common_value




    @exception_utils
    def write_excel(self):
        """运行结果写入excel"""
        l_reponse, l_ispass = read_txt_handel()
        print(l_reponse.__len__())
        print(l_ispass.__len__())

        i = 0
        j = 0
        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            # 实际结果列
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8, min_col=8):
                for cell in row:
                    cell.value = l_reponse[i]
                    # print("resp:%s" % i, cell.value)
                    i += 1
            # 是否通过列
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=9, min_col=9):
                for cell in row:
                    cell.value = l_ispass[j]
                    # print("ispass%s:" % j, cell.value)
                    j += 1

        save_path = "%s/output/run_result_excel/运行结果_%s.xlsx" % (base_dir, time.strftime("%Y%m%d_%H:%M:%S"))
        self.wb.save(save_path)

